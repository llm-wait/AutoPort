import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import ColorList

from Postman.tool.project_path import test_case_path
from Postman.tool.readconfig import ReadConfig

class ReadExcel():
    '''
    从配置文件case.conf中读取[MODE]->mode值,(测试的工作表和执行的模式)
    mode控制是否执行所有的用例，mode值只能是all或列表，默认值all表示执行所有的用例，
    列表表示执行指定的用例
    '''
    def __init__(self,filename):
       self.filename=filename
       self.wb=openpyxl.load_workbook(test_case_path+"/"+filename)
       self.confile="postmancase.conf"
    def  read_cases(self):
        #从配置文件中读取要测试的工作表和模式
        mode = eval(ReadConfig(self.confile).read_config("MODE", 'mode'))
        test_data=[]
        for key in mode:
            sheet=self.wb[key]#表单名
            if mode[key]=="all":
                for i in range(2,sheet.max_row+1):
                    row_data={}#单行测试用例
                    row_data["case_id"]=sheet.cell(i,1).value
                    row_data["module"]=sheet.cell(i,2).value
                    row_data["title"] = sheet.cell(i,3).value
                    row_data["url"] = sheet.cell(i,4).value
                    row_data["method"] = sheet.cell(i,5).value
                    row_data["header"] = sheet.cell(i,6).value
                    row_data["params"] = sheet.cell(i,7).value
                    row_data["expected"] = sheet.cell(i,8).value
                    row_data["sheetname"]=key#测试用例中添加sheetname字典，用于测试结果写回时之指定excel
                    row_data["depnedcase"]=sheet.cell(i,11).value
                    row_data["depnedvalue"] = sheet.cell(i, 12).value
                    row_data["casedepned"] = sheet.cell(i, 13).value
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}  # 单行测试用例
                    row_data["case_id"] = sheet.cell(case_id+1, 1).value
                    row_data["module"] = sheet.cell(case_id+1, 2).value
                    row_data["title"] = sheet.cell(case_id+1, 3).value
                    row_data["url"] = sheet.cell(case_id+1, 4).value
                    row_data["method"] = sheet.cell(case_id+1, 5).value
                    row_data["header"] = sheet.cell(case_id+1, 6).value
                    row_data["params"] = sheet.cell(case_id+1, 7).value
                    row_data["expected"] = sheet.cell(case_id+1, 8).value
                    row_data["sheetname"] = key
                    row_data["casedepnedname"] = sheet.cell(i, 11).value
                    row_data["casedepnedvalue"] = sheet.cell(i, 12).value
                    row_data["casedepned"] = sheet.cell(i, 13).value
                    test_data.append(row_data)
        return test_data
    def write_back(self,sheet_name,i,j,value,back_color=None):
        '''
        用于写回测试结果
        '''
        sheet=self.wb[sheet_name]
        sheet.cell(i,j).value=value
        #设置写回单元格填充颜色,默认是白色
        if back_color==None:
            sheet.cell(i,j).fill= PatternFill("solid",fgColor="FFFFFF")
        else:
            sheet.cell(i, j).fill = PatternFill("solid",fgColor=back_color)
        self.wb.save(test_case_path+"/"+self.filename)
    def read_depend_case(self,sheet_name,case_id):
        '''
        用于读取接口依赖数据
        根据传递的参数case_id,找到对应行的数据
        :return:
        '''
        sheet = self.wb[sheet_name]
        depend_case=[]
        for i in range (1,sheet.max_column+1):
            depend_case.append(sheet.cell(case_id+1,i).value)
        return depend_case

if __name__ == '__main__':
   filename=r"业务管理系统接口测试用例.xlsx"
   read=ReadExcel(filename)
   print(read.read_cases())
   red="FF0000"
   #read.write_back("登录测试用例",2,10,"aaaa")
   d=eval(read.read_depend_case("登录测试用例",9)[12])
   print(type(d))


